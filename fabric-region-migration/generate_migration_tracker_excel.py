#!/usr/bin/env python3
"""
generate_migration_tracker_excel.py

Generates a multi-sheet Excel workbook for tracking a Microsoft Fabric
capacity migration (e.g., West US → East US 2).

Sheets:
  1. Migration Inventory  – Placeholder for capacity/workspace/item data
                            (populated from the inventory notebook's CSV exports)
  2. Migration Phases     – 6-phase plan with milestones, dates, owners, status
  3. Risk Register        – Pre-populated migration risks with mitigations
  4. RACI Matrix          – Responsibility assignment for key migration tasks
  5. Issue Tracker        – Log for tracking issues during migration
  6. Validation Checklist – Post-migration validation steps per wave

Usage:
    pip install openpyxl
    python generate_migration_tracker_excel.py

Output:
    fabric_capacity_migration_tracker.xlsx
"""

from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "openpyxl is required.  Install with:  pip install openpyxl"
    )


# ── Style constants ─────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BORDER_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, num_cols):
    """Apply standard header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 30


def style_data_rows(ws, start_row, end_row, num_cols, alternate=True):
    """Apply borders & optional alternating fills to data rows."""
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = WRAP_ALIGN
        if alternate and (r - start_row) % 2 == 1:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL


def auto_width(ws, num_cols, min_width=12, max_width=50):
    """Set column widths based on content."""
    for col in range(1, num_cols + 1):
        header_len = len(str(ws.cell(row=1, column=col).value or ""))
        content_lens = [
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(2, ws.max_row + 1)
        ]
        best = max([header_len] + content_lens) + 4
        ws.column_dimensions[get_column_letter(col)].width = min(max(best, min_width), max_width)


# ── Sheet builders ──────────────────────────────────────────────────────────

def build_inventory_sheet(wb):
    """Sheet 1: Migration Inventory (template rows – fill from CSV)."""
    ws = wb.active
    ws.title = "Migration Inventory"

    headers = [
        "Workspace Name", "Capacity Name", "SKU", "Source Region",
        "Target Region", "Item Name", "Item Type",
        "Movable?", "Migration Complexity", "Suggested Wave",
        "Business Owner", "Data Size (GB)", "External Dependencies",
        "Notes"
    ]
    ws.append(headers)

    # Two example rows
    examples = [
        ["Sales Workspace", "FabCap-WestUS-01", "F64", "West US",
         "East US 2", "Sales Report", "Report",
         "✅ Yes", "Low", 1, "Jane Smith", "", "", ""],
        ["ETL Workspace", "FabCap-WestUS-01", "F64", "West US",
         "East US 2", "Ingest Pipeline", "DataPipeline",
         "❌ No", "High", 3, "Bob Jones", "~500",
         "Databricks (East US 2)", "Must re-create after reassignment"],
    ]
    for row in examples:
        ws.append(row)

    # 20 blank rows to fill
    for _ in range(20):
        ws.append([""] * len(headers))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"


def build_phases_sheet(wb):
    """Sheet 2: Migration Phases with milestones."""
    ws = wb.create_sheet("Migration Phases")

    headers = [
        "Phase", "Phase Name", "Key Milestone", "Owner",
        "Planned Start", "Planned End", "Actual Start", "Actual End",
        "Status", "% Complete", "Dependencies", "Notes"
    ]
    ws.append(headers)

    today = date.today()

    phases = [
        ("Phase 0", "Discovery & Inventory",
         "Complete inventory & get business sign-off",
         "", today, today + timedelta(weeks=2), "", "", "Not Started", "0%",
         "Admin API access", "Run migration-inventory notebook"),
        ("Phase 1", "Target Capacity Setup",
         "New capacity provisioned & validated in East US 2",
         "", today + timedelta(weeks=2), today + timedelta(weeks=3), "", "",
         "Not Started", "0%",
         "Azure subscription, budget approval",
         "Match SKU to source; configure networking"),
        ("Phase 2", "Wave 1 – Low Complexity",
         "All movable-only workspaces migrated & validated",
         "", today + timedelta(weeks=3), today + timedelta(weeks=5), "", "",
         "Not Started", "0%",
         "Phase 1 complete",
         "Reports, Dashboards, Semantic models"),
        ("Phase 3", "Wave 2 – Medium Complexity",
         "Lakehouses / Warehouses migrated with data copy",
         "", today + timedelta(weeks=5), today + timedelta(weeks=8), "", "",
         "Not Started", "0%",
         "Phase 2 complete, data copy tooling ready",
         "Use AzCopy / Copy Job for data transfer"),
        ("Phase 4", "Wave 3 – High Complexity",
         "Notebooks, Pipelines, Eventhouses recreated & validated",
         "", today + timedelta(weeks=8), today + timedelta(weeks=12), "", "",
         "Not Started", "0%",
         "Phase 3 complete, Git integration",
         "Pipeline re-creation; end-to-end testing"),
        ("Phase 5", "Validation, Cutover & Decommission",
         "Business sign-off, source decommissioned",
         "", today + timedelta(weeks=12), today + timedelta(weeks=14), "", "",
         "Not Started", "0%",
         "All phases complete",
         "Parallel run → cutover → decom"),
    ]

    for p in phases:
        ws.append(list(p))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

    # Status column validation hint (column 9)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")


def build_risk_register(wb):
    """Sheet 3: Risk Register with pre-populated migration risks."""
    ws = wb.create_sheet("Risk Register")

    headers = [
        "Risk ID", "Risk Description", "Category",
        "Likelihood (L/M/H)", "Impact (L/M/H)", "Risk Level",
        "Mitigation Strategy", "Risk Owner", "Status", "Notes"
    ]
    ws.append(headers)

    risks = [
        ("R-001",
         "Non-movable items block workspace reassignment",
         "Technical", "High", "High", "Critical",
         "Pre-inventory all items; remove non-movable items before reassignment",
         "", "Open", "Use inventory notebook to classify items"),
        ("R-002",
         "Dataflow Gen2 staging lakehouses block migration",
         "Technical", "Medium", "High", "High",
         "Delete all Dataflow Gen2 items first, then delete staging lakehouses",
         "", "Open", "MS Learn: staging items only visible after DFGen2 deletion"),
        ("R-003",
         "Admin API rate limits (200 req/hr) slow discovery",
         "Technical", "Medium", "Low", "Medium",
         "Implement retry/backoff (already built into notebook helper)",
         "", "Mitigated", "fabric_api_get() handles 429s automatically"),
        ("R-004",
         "Large-storage-format semantic models cannot cross regions",
         "Technical", "Medium", "High", "High",
         "Identify large-storage models early; switch to small storage or recreate",
         "", "Open", "Check semantic model storage format before migration"),
        ("R-005",
         "Data loss during Lakehouse/Warehouse migration",
         "Data", "Low", "Critical", "High",
         "Full backup before migration; validate row counts and checksums post-copy",
         "", "Open", "Use parallel copy + validation queries"),
        ("R-006",
         "Running jobs cancelled during workspace reassignment",
         "Operational", "High", "Medium", "High",
         "Schedule migration during maintenance windows; notify users in advance",
         "", "Open", "MS Learn: reassignment cancels all running jobs"),
        ("R-007",
         "Business disruption during migration window",
         "Business", "Medium", "High", "High",
         "Phased approach with parallel run; communicate schedule to stakeholders",
         "", "Open", "Wave-based migration reduces blast radius"),
        ("R-008",
         "External dependencies not updated (Databricks, gateways)",
         "Integration", "Medium", "High", "High",
         "Document all external deps in inventory; update connection strings post-migration",
         "", "Open", "Databricks already in East US 2 – update Fabric endpoints only"),
        ("R-009",
         "Private Link / VNet configuration breaks after migration",
         "Networking", "Medium", "High", "High",
         "Disable Private Link before migration; reconfigure in target region after",
         "", "Open", "MS Learn: Private Link must be temporarily disabled"),
        ("R-010",
         "Insufficient capacity SKU in target region",
         "Planning", "Low", "High", "Medium",
         "Verify target region SKU availability; request quota increase if needed",
         "", "Open", "Check Azure capacity quotas before Phase 1"),
    ]

    for r in risks:
        ws.append(list(r))

    # 10 blank rows
    for _ in range(10):
        ws.append([""] * len(headers))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"


def build_raci_matrix(wb):
    """Sheet 4: RACI Matrix for migration tasks."""
    ws = wb.create_sheet("RACI Matrix")

    headers = [
        "Task / Activity",
        "Fabric Admin", "Cloud Platform Team", "Data Engineering",
        "BI / Analytics", "Business Owner", "Security / Networking",
        "Project Manager"
    ]
    ws.append(headers)

    # R=Responsible, A=Accountable, C=Consulted, I=Informed
    tasks = [
        ("Run inventory notebook",               "R/A", "I",   "C",   "I",   "I",   "I",   "I"),
        ("Provision target Fabric capacity",      "R",   "A",   "I",   "I",   "I",   "C",   "I"),
        ("Configure networking (VNet/PE)",        "C",   "R/A", "I",   "I",   "I",   "R",   "I"),
        ("Backup lakehouse/warehouse data",       "C",   "I",   "R/A", "I",   "I",   "I",   "I"),
        ("Migrate Wave 1 workspaces (movable)",   "R/A", "I",   "I",   "C",   "I",   "I",   "I"),
        ("Copy data to target lakehouses",        "C",   "I",   "R/A", "I",   "I",   "I",   "I"),
        ("Remove non-movable items from source",  "R/A", "I",   "C",   "C",   "I",   "I",   "I"),
        ("Reassign workspaces to target capacity", "R/A", "I",   "I",   "I",   "I",   "I",   "I"),
        ("Recreate pipelines & notebooks",        "C",   "I",   "R/A", "I",   "I",   "I",   "I"),
        ("Recreate eventhouses & KQL databases",  "C",   "I",   "R/A", "I",   "I",   "I",   "I"),
        ("Validate reports & dashboards",         "I",   "I",   "I",   "R/A", "C",   "I",   "I"),
        ("Validate data pipelines end-to-end",    "I",   "I",   "R/A", "C",   "C",   "I",   "I"),
        ("Update Databricks connections",         "I",   "I",   "R/A", "I",   "I",   "I",   "I"),
        ("Business sign-off per wave",            "I",   "I",   "I",   "C",   "R/A", "I",   "I"),
        ("Decommission source capacities",        "R",   "A",   "I",   "I",   "I",   "C",   "I"),
        ("Update private endpoints / DNS",        "C",   "R",   "I",   "I",   "I",   "R/A", "I"),
        ("Overall project coordination",          "C",   "C",   "C",   "C",   "C",   "C",   "R/A"),
    ]

    for t in tasks:
        ws.append(list(t))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "B2"

    # Color-code RACI cells
    raci_fills = {
        "R":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
        "A":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # Blue
        "R/A": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),   # Bold green
        "C":   PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),   # Yellow
        "I":   PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),   # Light gray
    }
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            val = str(ws.cell(row=r, column=c).value or "").strip()
            if val in raci_fills:
                ws.cell(row=r, column=c).fill = raci_fills[val]
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")


def build_issue_tracker(wb):
    """Sheet 5: Issue Tracker."""
    ws = wb.create_sheet("Issue Tracker")

    headers = [
        "Issue ID", "Date Raised", "Workspace / Item",
        "Issue Description", "Severity (P1/P2/P3/P4)",
        "Assigned To", "Status", "Resolution", "Date Resolved", "Notes"
    ]
    ws.append(headers)

    # Example row
    ws.append([
        "ISS-001", "", "Example Workspace / Lakehouse",
        "(Example) Data mismatch after copy – row count differs by 5",
        "P2", "", "Open", "", "", ""
    ])

    # 30 blank rows
    for _ in range(30):
        ws.append([""] * len(headers))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"


def build_validation_checklist(wb):
    """Sheet 6: Post-Migration Validation Checklist."""
    ws = wb.create_sheet("Validation Checklist")

    headers = [
        "Wave", "Workspace Name", "Validation Step",
        "Expected Result", "Actual Result",
        "Pass/Fail", "Validated By", "Date", "Notes"
    ]
    ws.append(headers)

    checks = [
        (1, "(Wave 1 workspace)", "Reports render correctly",
         "All visuals load without error", "", "", "", "", ""),
        (1, "(Wave 1 workspace)", "Semantic model refresh succeeds",
         "Refresh completes < 2x baseline duration", "", "", "", "", ""),
        (1, "(Wave 1 workspace)", "Dashboard tiles load",
         "All tiles display data", "", "", "", "", ""),
        (1, "(Wave 1 workspace)", "Data gateway connectivity",
         "On-prem data sources accessible", "", "", "", "", ""),
        (2, "(Wave 2 workspace)", "Lakehouse row counts match source",
         "Row count delta = 0", "", "", "", "", ""),
        (2, "(Wave 2 workspace)", "Warehouse query results match",
         "Checksum/hash comparison passes", "", "", "", "", ""),
        (2, "(Wave 2 workspace)", "SQL endpoint accessible",
         "Queries execute successfully", "", "", "", "", ""),
        (2, "(Wave 2 workspace)", "Semantic models connect to new lakehouse",
         "No connection errors", "", "", "", "", ""),
        (3, "(Wave 3 workspace)", "Notebooks execute without error",
         "All cells pass", "", "", "", "", ""),
        (3, "(Wave 3 workspace)", "Data pipelines run end-to-end",
         "Pipeline succeeds with expected output", "", "", "", "", ""),
        (3, "(Wave 3 workspace)", "Eventhouse ingestion active",
         "Events streaming into KQL database", "", "", "", "", ""),
        (3, "(Wave 3 workspace)", "Spark jobs complete on schedule",
         "Job duration within 2x baseline", "", "", "", "", ""),
        (3, "(Wave 3 workspace)", "Databricks connections functional",
         "Fabric pipeline invokes Databricks successfully", "", "", "", "", ""),
        ("All", "ALL", "Capacity utilization within thresholds",
         "CU% < 80% sustained", "", "", "", "", ""),
        ("All", "ALL", "No orphaned items in source region",
         "Source capacity empty", "", "", "", "", ""),
    ]

    for c in checks:
        ws.append(list(c))

    # 15 blank rows
    for _ in range(15):
        ws.append([""] * len(headers))

    style_header_row(ws, len(headers))
    style_data_rows(ws, 2, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    build_inventory_sheet(wb)
    build_phases_sheet(wb)
    build_risk_register(wb)
    build_raci_matrix(wb)
    build_issue_tracker(wb)
    build_validation_checklist(wb)

    output = "fabric_capacity_migration_tracker.xlsx"
    wb.save(output)
    print(f"✅ Migration tracker generated: {output}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    print()
    print("Next steps:")
    print("  1. Run the inventory notebook to generate CSV exports")
    print("  2. Paste CSV data into the 'Migration Inventory' sheet")
    print("  3. Fill in owners, dates, and external dependencies")
    print("  4. Share with stakeholders for review")


if __name__ == "__main__":
    main()
